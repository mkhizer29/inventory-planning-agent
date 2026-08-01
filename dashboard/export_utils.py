"""dashboard/export_utils.py — one reusable table-export component for the dashboard.

Every user-facing table gets a single compact ``Export ▾`` menu instead of three
big buttons. All three formats are built **in memory** (BytesIO) and handed to
``st.download_button`` — nothing is ever written to outputs/, runs/, data/ or the
repository root.

Pure builders (``dataframe_to_*_bytes``) contain no Streamlit imports at module
scope, so they are unit-testable without a Streamlit runtime.
"""
from __future__ import annotations

import io
import re
from datetime import datetime
from zoneinfo import ZoneInfo

import pandas as pd

DISPLAY_TIMEZONE = ZoneInfo("Asia/Karachi")
MAX_PDF_COLUMNS = 10          # wider tables export user-facing columns only (CSV/Excel stay complete)


# ── shared helpers ────────────────────────────────────────────────────────────────────
def _now_pkt() -> str:
    return datetime.now(DISPLAY_TIMEZONE).strftime("%d %b %Y · %I:%M %p PKT")


def safe_filename(stem: str, ext: str) -> str:
    """A filesystem-safe download name (the browser names the file; we never write it)."""
    clean = re.sub(r"[^A-Za-z0-9._-]+", "_", str(stem)).strip("_") or "export"
    stamp = datetime.now(DISPLAY_TIMEZONE).strftime("%Y%m%d_%H%M")
    return f"{clean}_{stamp}.{ext}"


def _as_frame(df) -> pd.DataFrame:
    if df is None:
        return pd.DataFrame()
    return df if isinstance(df, pd.DataFrame) else pd.DataFrame(df)


def _is_null(v) -> bool:
    """True for None / NaN / NaT (pd.NaT is a datetime subclass, so it must be caught first)."""
    if v is None or v is pd.NaT:
        return True
    try:
        return bool(pd.isna(v))
    except (TypeError, ValueError):
        return False


def _fmt_cell(v) -> str:
    """Readable text for PDF cells; preserves the underlying value's meaning.
    Null stays blank (never coerced to 0) and long text is returned untruncated."""
    if _is_null(v):
        return ""
    if isinstance(v, float):
        return f"{v:,.2f}" if abs(v) < 1e12 else f"{v:.3g}"
    if isinstance(v, (int,)) and not isinstance(v, bool):
        return f"{v:,}"
    if isinstance(v, (pd.Timestamp, datetime)):
        return v.strftime("%Y-%m-%d")
    return str(v)


# ── CSV ────────────────────────────────────────────────────────────────────────────────
def dataframe_to_csv_bytes(df) -> bytes:
    """UTF-8 **with BOM** so Excel opens non-ASCII product names correctly.
    Exports every row given (i.e. the caller's currently filtered view), full values,
    stable column order."""
    return _as_frame(df).to_csv(index=False).encode("utf-8-sig")


# ── Excel ──────────────────────────────────────────────────────────────────────────────
def dataframe_to_excel_bytes(df, *, sheet_name: str = "Data", title: str | None = None,
                             metadata: dict | None = None) -> bytes:
    """.xlsx via openpyxl: bold + frozen header, autofilter, sized columns, wrapped text,
    and an optional title/metadata block above the table."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    frame = _as_frame(df)
    safe_sheet = re.sub(r"[\[\]\*\?/\\:]", "-", str(sheet_name))[:31] or "Data"
    wb = Workbook()
    ws = wb.active
    ws.title = safe_sheet

    row_cursor = 1
    if title:
        ws.cell(row=1, column=1, value=str(title)).font = Font(bold=True, size=14)
        row_cursor = 2
    if metadata:
        for k, v in metadata.items():
            ws.cell(row=row_cursor, column=1, value=str(k)).font = Font(bold=True, size=9)
            ws.cell(row=row_cursor, column=2, value=str(v)).font = Font(size=9)
            row_cursor += 1
    if title or metadata:
        row_cursor += 1                                     # spacer before the header

    header_row = row_cursor
    head_fill = PatternFill("solid", fgColor="0B1F33")      # Naheed navy
    for j, col in enumerate(frame.columns, start=1):
        c = ws.cell(row=header_row, column=j, value=str(col))
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = head_fill
        c.alignment = Alignment(vertical="center", wrap_text=True)

    for i, (_, row) in enumerate(frame.iterrows(), start=header_row + 1):
        for j, col in enumerate(frame.columns, start=1):
            v = row[col]
            if _is_null(v):
                v = None                                    # keep nulls empty, never 0 (NaT included)
            elif isinstance(v, (pd.Timestamp, datetime)):
                v = v.strftime("%Y-%m-%d")
            elif not isinstance(v, (int, float, str, bool)):
                v = str(v)
            ws.cell(row=i, column=j, value=v).alignment = Alignment(wrap_text=True, vertical="top")

    n_rows = len(frame)
    if len(frame.columns):
        ws.freeze_panes = ws.cell(row=header_row + 1, column=1)
        ws.auto_filter.ref = (f"A{header_row}:"
                              f"{get_column_letter(len(frame.columns))}{header_row + max(n_rows, 1)}")
        for j, col in enumerate(frame.columns, start=1):
            longest = max([len(str(col))] + [len(_fmt_cell(v)) for v in frame[col].head(200)] or [0])
            ws.column_dimensions[get_column_letter(j)].width = min(max(longest + 2, 10), 48)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── PDF ────────────────────────────────────────────────────────────────────────────────
def dataframe_to_pdf_bytes(df, *, title: str, metadata: dict | None = None,
                           landscape: bool = True) -> bytes:
    """A readable A4 table report via reportlab: repeated header row, page numbers,
    PKT generation timestamp, run/filter metadata, and wrapped long product names.

    Very wide tables export the first ``MAX_PDF_COLUMNS`` user-facing columns and say so —
    CSV/Excel remain the complete dataset.
    """
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape as _landscape
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import (BaseDocTemplate, Frame, PageTemplate, Paragraph,
                                    Spacer, Table, TableStyle)

    frame = _as_frame(df)
    all_cols = list(frame.columns)
    cols = all_cols[:MAX_PDF_COLUMNS]
    omitted = all_cols[len(cols):]

    pagesize = _landscape(A4) if landscape else A4
    buf = io.BytesIO()
    doc = BaseDocTemplate(buf, pagesize=pagesize, leftMargin=12 * mm, rightMargin=12 * mm,
                          topMargin=12 * mm, bottomMargin=14 * mm, title=str(title))
    frame_area = Frame(doc.leftMargin, doc.bottomMargin, doc.width, doc.height, id="body")

    def _page_furniture(canvas, doc_):
        canvas.saveState()
        canvas.setFont("Helvetica", 7.5)
        canvas.setFillColor(colors.HexColor("#6B7A8F"))
        canvas.drawString(doc_.leftMargin, 8 * mm, "Naheed · Inventory Planning Agent")
        canvas.drawRightString(pagesize[0] - doc_.rightMargin, 8 * mm, f"Page {canvas.getPageNumber()}")
        canvas.restoreState()

    doc.addPageTemplates([PageTemplate(id="main", frames=[frame_area], onPage=_page_furniture)])

    ss = getSampleStyleSheet()
    h_style = ParagraphStyle("h", parent=ss["Title"], fontSize=15, spaceAfter=4,
                             textColor=colors.HexColor("#0B1F33"), alignment=0)
    meta_style = ParagraphStyle("m", parent=ss["Normal"], fontSize=8,
                                textColor=colors.HexColor("#6B7A8F"), leading=11)
    cell_style = ParagraphStyle("c", parent=ss["Normal"], fontSize=7.2, leading=8.8)
    head_style = ParagraphStyle("hd", parent=ss["Normal"], fontSize=7.4, leading=9,
                                textColor=colors.white, fontName="Helvetica-Bold")

    story = [Paragraph(str(title), h_style),
             Paragraph(f"Generated {_now_pkt()} · {len(frame):,} rows", meta_style)]
    if metadata:
        story.append(Paragraph(" · ".join(f"<b>{k}:</b> {v}" for k, v in metadata.items()), meta_style))
    if omitted:
        story.append(Paragraph(
            f"Showing {len(cols)} of {len(all_cols)} columns. Omitted: {', '.join(map(str, omitted))}. "
            "Use the CSV or Excel export for the complete dataset.", meta_style))
    story.append(Spacer(1, 6))

    if cols:
        data = [[Paragraph(str(c), head_style) for c in cols]]
        for _, row in frame.iterrows():
            data.append([Paragraph(_fmt_cell(row[c]).replace("&", "&amp;")
                                   .replace("<", "&lt;").replace(">", "&gt;"), cell_style)
                         for c in cols])
        usable = doc.width
        table = Table(data, colWidths=[usable / len(cols)] * len(cols), repeatRows=1)
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0B1F33")),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#E7ECF2")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F7F9FC")]),
            ("TOPPADDING", (0, 0), (-1, -1), 3), ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            ("LEFTPADDING", (0, 0), (-1, -1), 4), ("RIGHTPADDING", (0, 0), (-1, -1), 4),
        ]))
        story.append(table)
    else:
        story.append(Paragraph("No rows to export.", meta_style))

    doc.build(story)
    return buf.getvalue()


# ── the reusable Streamlit menu ────────────────────────────────────────────────────────
def render_table_export_menu(df, *, filename_stem: str, title: str,
                             metadata: dict | None = None, key: str,
                             landscape: bool = True) -> None:
    """One compact ``Export ▾`` control (popover; expander fallback) with CSV / Excel / PDF.

    Always exports the *currently filtered* frame handed in. Buffers are built lazily so a
    heavy PDF is only rendered when the menu is opened.
    """
    import streamlit as st

    frame = _as_frame(df)
    container = (st.popover("Export", use_container_width=False)
                 if hasattr(st, "popover") else st.expander("Export"))
    with container:
        if frame.empty:
            st.caption("Nothing to export in the current view.")
            return
        st.caption(f"{len(frame):,} rows · current filters")
        st.download_button("Download CSV", data=dataframe_to_csv_bytes(frame),
                           file_name=safe_filename(filename_stem, "csv"),
                           mime="text/csv", key=f"{key}_csv", use_container_width=True)
        try:
            st.download_button("Download Excel", data=dataframe_to_excel_bytes(
                                   frame, sheet_name=filename_stem[:31], title=title, metadata=metadata),
                               file_name=safe_filename(filename_stem, "xlsx"),
                               mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                               key=f"{key}_xlsx", use_container_width=True)
        except Exception as exc:                                    # noqa: BLE001
            st.caption(f"Excel export unavailable: {exc}")
        try:
            st.download_button("Download PDF", data=dataframe_to_pdf_bytes(
                                   frame, title=title, metadata=metadata, landscape=landscape),
                               file_name=safe_filename(filename_stem, "pdf"),
                               mime="application/pdf", key=f"{key}_pdf", use_container_width=True)
        except Exception as exc:                                    # noqa: BLE001
            st.caption(f"PDF export unavailable: {exc}")
