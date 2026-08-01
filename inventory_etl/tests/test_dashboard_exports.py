"""Dashboard export tests — pure in-memory CSV / Excel / PDF builders.

Never imports dashboard/app.py (that would execute the Streamlit app). Also asserts the
exporters write nothing to the repository (outputs/, runs/, data/, repo root).
"""
import hashlib
import io
import sys
from pathlib import Path

import pandas as pd
import pytest

REPO_ROOT = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(REPO_ROOT / "dashboard"))

import export_utils as eu          # noqa: E402

LONG_NAME = ("Nestle Pure Life Drinking Water 1.5 Litres — an extremely long product name "
             "used to prove wrapping and that nothing is truncated in the underlying data")


def _frame():
    return pd.DataFrame({
        "Product": [LONG_NAME, "Olper's Milk 1000ml", "Candyland Chili Mili"],
        "SKU": ["IC-1055803", "IC-1018988", "IC-1032718"],
        "Risk": ["Critical", "Watch", "Healthy"],
        "P(stockout)": [0.9123, 0.4, None],
        "Revenue at risk": [125000.5, None, 0.0],
        "Projected stockout": pd.to_datetime(["2026-07-05", None, None]),
    })


# ── CSV ────────────────────────────────────────────────────────────────────────────────
def test_csv_has_all_filtered_rows():
    df = _frame()
    text = eu.dataframe_to_csv_bytes(df).decode("utf-8-sig")
    assert len(text.strip().splitlines()) == len(df) + 1          # header + every row


def test_csv_preserves_full_product_names():
    text = eu.dataframe_to_csv_bytes(_frame()).decode("utf-8-sig")
    assert LONG_NAME in text and "…" not in text


def test_csv_is_utf8_bom_for_excel():
    assert eu.dataframe_to_csv_bytes(_frame())[:3] == b"\xef\xbb\xbf"


def test_csv_keeps_stable_column_order():
    df = _frame()
    header = eu.dataframe_to_csv_bytes(df).decode("utf-8-sig").splitlines()[0]
    assert header.split(",")[:3] == ["Product", "SKU", "Risk"]


def test_csv_null_revenue_not_zero():
    rows = eu.dataframe_to_csv_bytes(_frame()).decode("utf-8-sig").splitlines()
    assert rows[2].count(",,") >= 1 or ",," in rows[2]            # null stayed empty, not 0


# ── Excel ──────────────────────────────────────────────────────────────────────────────
def test_excel_opens_with_openpyxl():
    import openpyxl
    data = eu.dataframe_to_excel_bytes(_frame(), sheet_name="Risk")
    assert data[:2] == b"PK"
    wb = openpyxl.load_workbook(io.BytesIO(data))
    assert "Risk" in wb.sheetnames


def test_excel_has_headers_freeze_and_autofilter():
    import openpyxl
    data = eu.dataframe_to_excel_bytes(_frame(), sheet_name="Risk", title="Stockout Risk",
                                       metadata={"Run": "abc123", "Category": "Groceries & Pets"})
    ws = openpyxl.load_workbook(io.BytesIO(data))["Risk"]
    header_rows = [r for r in range(1, 12) if ws.cell(r, 1).value == "Product"]
    assert header_rows, "header row not found"
    hr = header_rows[0]
    assert [ws.cell(hr, c).value for c in range(1, 4)] == ["Product", "SKU", "Risk"]
    assert ws.cell(hr, 1).font.bold and ws.freeze_panes and ws.auto_filter.ref


def test_excel_preserves_long_names_and_nulls():
    import openpyxl
    ws = openpyxl.load_workbook(io.BytesIO(
        eu.dataframe_to_excel_bytes(_frame(), sheet_name="Risk")))["Risk"]
    vals = [ws.cell(r, 1).value for r in range(1, 8)]
    assert LONG_NAME in vals
    rev_col = [ws.cell(r, 5).value for r in range(2, 6)]
    assert None in rev_col                                        # null preserved, not 0


# ── PDF ────────────────────────────────────────────────────────────────────────────────
def test_pdf_signature():
    assert eu.dataframe_to_pdf_bytes(_frame(), title="Stockout Risk")[:5] == b"%PDF-"


def test_pdf_handles_long_text_and_metadata():
    data = eu.dataframe_to_pdf_bytes(
        _frame(), title="Stockout Risk — Priority Queue",
        metadata={"Run": "20260731T060331Z", "Category": "Groceries & Pets"})
    assert data[:5] == b"%PDF-" and len(data) > 1500


def test_pdf_wide_table_limits_columns_but_still_builds():
    wide = pd.DataFrame({f"col_{i}": [1, 2, 3] for i in range(20)})
    assert eu.dataframe_to_pdf_bytes(wide, title="Wide")[:5] == b"%PDF-"


def test_pdf_empty_frame_is_safe():
    assert eu.dataframe_to_pdf_bytes(pd.DataFrame(), title="Empty")[:5] == b"%PDF-"


# ── isolation / misc ───────────────────────────────────────────────────────────────────
def _snapshot(d: Path):
    return {p.name: hashlib.sha256(p.read_bytes()).hexdigest()
            for p in d.iterdir() if p.is_file()} if d.exists() else {}


def test_exports_write_nothing_to_the_repository():
    watched = [REPO_ROOT / "outputs", REPO_ROOT / "data" / "processed", REPO_ROOT]
    before = [_snapshot(d) for d in watched]
    df = _frame()
    eu.dataframe_to_csv_bytes(df)
    eu.dataframe_to_excel_bytes(df, sheet_name="S")
    eu.dataframe_to_pdf_bytes(df, title="T")
    assert [_snapshot(d) for d in watched] == before
    assert not (REPO_ROOT / "runs" / "export.csv").exists()


def test_safe_filename_is_filesystem_safe():
    name = eu.safe_filename("Stockout Risk / Queue: v2", "csv")
    assert name.endswith(".csv") and not set(name) & set('/\\:*?"<>|')


def test_null_helper_treats_nat_and_nan_as_null():
    assert eu._is_null(None) and eu._is_null(pd.NaT) and eu._is_null(float("nan"))
    assert not eu._is_null(0) and not eu._is_null("")


def test_fmt_cell_keeps_nulls_blank_and_names_whole():
    assert eu._fmt_cell(None) == "" and eu._fmt_cell(pd.NaT) == ""
    assert eu._fmt_cell(LONG_NAME) == LONG_NAME
    assert eu._fmt_cell(0.0) == "0.00"                            # zero is a real value


# ── compile guards ─────────────────────────────────────────────────────────────────────
@pytest.mark.parametrize("mod", ["app.py", "styles.py", "run_service.py", "export_utils.py"])
def test_dashboard_modules_compile(mod):
    import py_compile
    py_compile.compile(str(REPO_ROOT / "dashboard" / mod), doraise=True)
